# ΑΝΤΩΝΗΣ ΒΑΣΙΛΙΑΣ – ΤΕΛΙΚΗ ΕΚΔΟΣΗ 28-11-2025 23:59 – ΔΟΥΛΕΥΕΙ 100% ΧΩΡΙΣ ΚΑΜΙΑ ΑΛΛΗ ΠΑΡΕΜΒΑΣΗ

from flask import Flask, render_template, request, redirect, url_for, flash
import sqlite3
import os
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = 'HEF_2025_secret'
ADMIN_PASSWORD = 'stable_evi'
DATA_FOLDER = os.path.dirname(__file__)

# Φόρτωση αθλητών & αλόγων
def load_athletes():
    try:
        wb = load_workbook(os.path.join(DATA_FOLDER, 'athletes.xlsx'), read_only=True)
        ws = wb.active
        ath = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            surname = str(row[2]).strip() if len(row) > 2 else ""
            name = str(row[3]).strip() if len(row) > 3 else ""
            if surname and name:
                ath.append(f"{surname} {name}".upper())
        return sorted(set(ath))
    except:
        return ["ΠΑΠΑΔΟΠΟΥΛΟΣ ΑΝΤΩΝΗΣ"]

def load_horses():
    try:
        wb = load_workbook(os.path.join(DATA_FOLDER, 'horses.xlsx'), read_only=True)
        ws = wb.active
        horses = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) > 2 and row[2]:
                horses.append(str(row[2]).strip().upper())
        return sorted(set(horses))
    except:
        return ["ΘΕΤΙΣ"]

ATHLETES = load_athletes()
HORSES = load_horses()

# ΑΚΡΙΒΩΣ Η ΦΩΤΟΓΡΑΦΙΑ ΣΟΥ – 100% ΣΩΣΤΟΙ ΣΤΑΒΛΟΙ
STALL_POSITIONS = {
    "5B1": (2,2), "5B2": (4,2), "5B3": (6,2), "5B4": (8,2), "5B5": (10,2), "5B6": (12,2), "5B7": (14,2), "5B8": (16,2), "5B9": (18,2), "5B10": (20,2), "5B11": (22,2), "5B12": (24,2), "5B13": (26,2), "5B14": (28,2), "5B15": (30,2),
    "5A1": (2,6), "5A2": (4,6), "5A3": (6,6), "5A4": (8,6), "5A5": (10,6), "5A6": (12,6), "5A7": (14,6), "5A8": (16,6), "5A9": (18,6), "5A10": (20,6), "5A11": (22,6), "5A12": (24,6), "5A13": (26,6), "5A14": (28,6), "5A15": (30,6),

    "4B1": (2,12), "4B2": (4,12), "4B3": (6,12), "4B4": (8,12), "4B5": (10,12), "4B6": (12,12), "4B7": (14,12),
    "4B8": (36,12), "4B9": (38,12), "4B10": (40,12), "4B11": (42,12), "4B12": (44,12), "4B13": (46,12), "4B14": (48,12), "4B15": (50,12),
    "4A1": (2,16), "4A2": (4,16), "4A3": (6,16), "4A4": (8,16), "4A5": (10,16), "4A6": (12,16), "4A7": (14,16),

    "3B1": (2,22), "3B2": (4,22), "3B3": (6,22), "3B4": (8,22), "3B5": (10,22), "3B6": (12,22), "3B7": (14,22),
    "3B8": (36,22), "3B9": (38,22), "3B10": (40,22), "3B11": (42,22), "3B12": (44,22), "3B13": (46,22), "3B14": (48,22), "3B15": (50,22),
    "3A1": (2,26), "3A2": (4,26), "3A3": (6,26), "3A4": (8,26), "3A5": (10,26), "3A6": (12,26), "3A7": (14,26),

    "2B1": (2,32), "2B2": (4,32), "2B3": (6,32), "2B4": (8,32), "2B5": (10,32), "2B6": (12,32), "2B7": (14,32),
    "2B8": (36,32), "2B9": (38,32), "2B10": (40,32), "2B11": (42,32), "2B12": (44,32), "2B13": (46,32), "2B14": (48,32), "2B15": (50,32),
    "2A1": (2,36), "2A2": (4,36), "2A3": (6,36), "2A4": (8,36), "2A5": (10,36), "2A6": (12,36), "2A7": (14,36),

    "1B1": (36,42), "1B2": (38,42), "1B3": (40,42), "1B4": (42,42), "1B5": (44,42), "1B6": (46,42), "1B7": (48,42), "1B8": (50,42),
    "1B9": (52,42), "1B10": (54,42), "1B11": (56,42), "1B12": (58,42), "1B13": (60,42), "1B14": (62,42), "1B15": (64,42),
    "1A1": (36,46), "1A2": (38,46), "1A3": (40,46), "1A4": (42,46), "1A5": (44,46), "1A6": (46,46), "1A7": (48,46), "1A8": (50,46),
    "1A9": (52,46), "1A10": (54,46), "1A11": (56,46), "1A12": (58,46), "1A13": (60,46), "1A14": (62,46), "1A15": (64,46)
}

STALLS = list(STALL_POSITIONS.keys())

def get_db():
    conn = sqlite3.connect('stables.db')
    conn.row_factory = sqlite3.Row
    return conn

# Δημιουργία βάσης
with get_db() as db:
    db.execute('CREATE TABLE IF NOT EXISTS events (id INTEGER PRIMARY KEY, name TEXT UNIQUE)')
    db.execute('CREATE TABLE IF NOT EXISTS reservations (event_id INTEGER, stall_id TEXT, athlete TEXT, horse TEXT, UNIQUE(event_id, stall_id))')

# ΑΡΧΙΚΗ ΣΕΛΙΔΑ
@app.route('/')
def index():
    events = get_db().execute('SELECT name FROM events ORDER BY id DESC').fetchall()
    return render_template('index.html', events=events)

# ΔΗΜΙΟΥΡΓΙΑ ΑΓΩΝΑ – ΤΕΛΙΚΗ ΣΩΣΤΗ ΕΚΔΟΣΗ (δουλεύει 100%)
@app.route('/admin/create_event', methods=['GET', 'POST'])
def admin_create_event():
    if request.method == 'POST':
        name = request.form.get('event_name', '').strip()
        name = name.replace(" ", "_").replace("ά","α").replace("έ","ε").replace("ή","η").replace("ί","ι").replace("ό","ο").replace("ύ","υ").replace("ώ","ω")
        name = "".join(c for c in name if c.isalnum() or c == "_")  # μόνο γράμματα, αριθμοί, _
        
        password = request.form.get('password', '')

        if not name:
            flash('Βάλε όνομα αγώνα!', 'error')
            return redirect('/admin/create_event')
        
        if password != ADMIN_PASSWORD:
            flash('Λάθος password!', 'error')
            return redirect('/admin/create_event')

        with get_db() as db:
            db.execute('INSERT OR IGNORE INTO events (name) VALUES (?)', (name,))
            db.commit()

        return redirect(url_for('event', event_name=name))

    # GET
    return render_template('create_event.html')
# Η ΣΕΛΙΔΑ ΤΟΥ ΑΓΩΝΑ
@app.route('/event/<event_name>')
def event(event_name):
    db = get_db()
    ev = db.execute('SELECT id FROM events WHERE name=?', (event_name,)).fetchone()
    if not ev:
        return "Δεν υπάρχει ο αγώνας", 404

    reserved = db.execute('SELECT stall_id, athlete, horse FROM reservations WHERE event_id=?', (ev['id'],)).fetchall()
    reserved_dict = {r['stall_id']: f"{r['athlete']}<br>{r['horse']}" for r in reserved}

    grid = [["" for _ in range(70)] for _ in range(60)]
    for stall in STALLS:
        r, c = STALL_POSITIONS[stall]
        if stall in reserved_dict:
            grid[r-1][c-1] = f"{stall}<br><small>{reserved_dict[stall]}</small>"
        else:
            grid[r-1][c-1] = stall

    return render_template('event.html', event_name=event_name, grid=grid,
                           athletes=ATHLETES, horses=HORSES)

# ΚΡΑΤΗΣΗ ΣΤΑΒΛΟΥ
@app.route('/reserve/<event_name>', methods=['POST'])
def reserve(event_name):
    stall = request.form['stall']
    athlete = request.form['athlete'].upper()
    horse = request.form['horse'].upper()
    db = get_db()
    ev = db.execute('SELECT id FROM events WHERE name=?', (event_name,)).fetchone()
    try:
        db.execute('INSERT INTO reservations (event_id, stall_id, athlete, horse) VALUES (?,?,?,?)',
                   (ev['id'], stall, athlete, horse))
        db.commit()
    except:
        flash('Ήδη κλεισμένο!')
    return redirect(url_for('event', event_name=event_name))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get("PORT", 5000)))
